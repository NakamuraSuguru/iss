import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
import math
import unicodedata
import pandas as pd

def main():
    #計算結果のファイルを設定
    input_file = 'output.txt'
    #勤務表のファイル名を設定
    output_file = 'iss_result.xlsx'
    staffs,days,slots,positions,opt_list,salarys,dw = read_sol(input_file)
    write_schedule(staffs,days,slots,salarys,dw,positions,opt_list,output_file)
    return

thin = Side(style='thin', color='000000')
thin_red = Side(style='thin', color='ff0000')
double = Side(style='double', color='000000')

#1行目の値からcharaの列を返す
def found_col(sheet_name,chara):
    for col in sheet_name.iter_cols():
        if col[0].value == chara:
            column = col[0].column
            break
    return column

#1列目の値からcharaの行を返す
def found_row(sheet_name,chara):
    for row in sheet_name.iter_rows():
        if row[0].value == chara:
            row2 = row[0].row
            break
    return row2

#計算結果のファイルを読み込む
def read_sol(file_name):
    f = open(file_name,'r',encoding='utf-8')

    result_list = f.readlines()

    idx_staff = result_list.index('---スタッフ---\n')
    staffs = result_list[idx_staff+1].rstrip('\n').split(',')

    idx_day = result_list.index('---日数---\n')
    days = result_list[idx_day+1].rstrip('\n').split(',')

    idx_open = result_list.index('---オープン---\n')
    open_time = float(result_list[idx_open+1].rstrip('\n'))

    idx_close = result_list.index('---クローズ---\n')
    close_time = float(result_list[idx_close+1].rstrip('\n'))

    slots = make_slot(open_time,close_time)

    idx_position = result_list.index('---ポジション---\n')
    positions = result_list[idx_position+1].rstrip('\n').split(';')

    idx_opt = result_list.index('---解---\n')
    opt_list,staffs2 = make_opt_list(result_list,idx_opt,staffs,positions,slots,days)

    idx_salary = result_list.index('---各スタッフの給与---\n')
    salarys = make_staff_salary(result_list,idx_salary,staffs,staffs2)

    idx_dw = result_list.index('---各スタッフの希望総給与---\n')
    dw = make_staff_salary(result_list,idx_dw,staffs,staffs2)

    f.close()
    return staffs,days,slots,positions,opt_list,salarys,dw
    
#時間帯のリストを作成
def make_slot(open_time,close_time):
    slot = []
    interval = int((close_time - open_time) * 2)
    for i in range(interval):
        slot.append(str(open_time+i*0.5))
    slot.append(str(close_time))
    return slot

#各スタッフの給与を読み込む
def make_staff_salary(result_list,idx_salary,staffs,staffs2):
    staff_salary = {}
    for i in range(len(staffs2)):
        r_l = result_list[idx_salary+1+i].rstrip('\n').split(':')
        if r_l[0] == staffs2[i]:
            staff_salary[r_l[0]] = int(r_l[1])
        else:
            staff_salary[staffs[i]] = 0
    return staff_salary

#得られた解をリスト化
def make_opt_list(result_list,idx_opt,staffs,positions,slots,days):
    df_dict = {}
    max_h = 1
    max_p = 1
    for i in staffs:
        df_dict[i] = pd.DataFrame(index=slots, columns=days)
    for k in result_list[idx_opt+1:]:
        if k.rstrip('\n').split(';')[0] == '不足':
            for h in range(1,max_h+1):
                if '不足'+str(h) not in list(df_dict.keys()):
                    staffs.append('不足'+str(h))
                    df_dict['不足'+str(h)] = pd.DataFrame(index=slots, columns=days)
                if df_dict['不足'+str(h)].isnull()[str(k.rstrip('\n').split(';')[1])][k.rstrip('\n').split(';')[2]]:
                    df_dict['不足'+str(h)][str(k.rstrip('\n').split(';')[1])][k.rstrip('\n').split(';')[2]] = k.rstrip('\n').split(';')[3]
                    max_h += 1
                    break
                else:
                    continue
        else:
            df_dict[k.rstrip('\n').split(';')[0]][str(k.rstrip('\n').split(';')[1])][k.rstrip('\n').split(';')[2]] = k.rstrip('\n').split(';')[3]

    for j in days:
        df_dict[str(j)+'日'] = pd.DataFrame(index=slots, columns=staffs+positions)
        for i in staffs:
            df_dict[str(j)+'日'][i] = df_dict[i][j]
            for m in slots:
                if df_dict[i].notnull()[j][m]:
                    if df_dict[str(j)+'日'].isnull()[df_dict[i][j][m]][m]:
                        df_dict[str(j)+'日'][df_dict[i][j][m]][m] = i
                    else:
                        for p in range(1,max_p+1):
                            if df_dict[i][j][m] + str(p) not in list(df_dict[str(j)+'日'].keys()):
                                s = pd.Series(index=slots)
                                df_dict[str(j)+'日'].insert(list(df_dict[str(j)+'日'].keys()).index(df_dict[i][j][m])+p,df_dict[i][j][m] + str(p),s)
                                df_dict[str(j)+'日'][df_dict[i][j][m] + str(p)][m] = i
                                max_p += 1
                                break
                            else:
                                if df_dict[str(j)+'日'].isnull()[df_dict[i][j][m] + str(p)][m]:
                                    df_dict[str(j)+'日'][df_dict[i][j][m] + str(p)][m] = i
                                    break
    return df_dict,staffs

#エクセルファイルへ書き込む
def write_schedule(staffs,days,slots,salarys,dw,positions,opt_list,file_name):
    wb = openpyxl.Workbook()

    sheet = wb['Sheet']
    sheet.title = '月間シフト'

    #月間シフト作成
    write_monthly(sheet,staffs,days,opt_list)
    days2 = [d + '日' for d in days]
    change_h_axis(sheet,days2,1)
    adjust_width(sheet)

    #日別シフト作成
    slots2 = [slot.replace('.0',':00')+'～' if '.0' in slot else slot.replace('.5',':30')+'～' for slot in slots]
    write_daily(wb,opt_list,days,staffs,slots2)

    #スタッフ別シフト作成
    write_staff(wb,opt_list,staffs,days,positions,slots2)

    #詳細データ
    detail = wb.create_sheet(title='詳細')
    h_axis = ['総給与','希望総給与','希望総給与との差(絶対値)','総出勤日数','休み日数','総勤務時間','夜勤以外の勤務時間','夜勤の勤務時間','最大連続勤務日数']
    positions_num = [p+'の勤務日数' for p in positions]
    h_axis.extend(positions_num)
    make_axis(detail,staffs+['合計'],h_axis,'詳細')

    #希望総給与
    write_detail(detail,dw,'希望総給与')

    #出勤数,休み数,最大連続勤務
    sum_work_day,sum_off,max_conti = count_work_day(sheet,staffs)
    write_detail(detail,sum_work_day,'総出勤日数')
    write_detail(detail,sum_off,'休み日数')
    write_detail(detail,max_conti,'最大連続勤務日数')

    #総給与，勤務時間，夜勤勤務時間，それ以外の勤務，ポジション回数
    sum_work_time,sum_day,sum_night,sum_salary,staffs_posi = count_work_time(opt_list,days,staffs,positions,salarys)
    write_detail(detail,sum_work_time,'総勤務時間')
    write_detail(detail,sum_day,'夜勤以外の勤務時間')
    write_detail(detail,sum_night,'夜勤の勤務時間')
    write_detail(detail,sum_salary,'総給与')
    for p in positions:
        write_detail(detail,staffs_posi[p],p+'の勤務日数')

    #希望総給与との差
    abs_sub = dict(zip(staffs,[0]*len(staffs)))
    for i in staffs:
        abs_sub[i] = abs(sum_salary[i]-dw[i])
    write_detail(detail,abs_sub,'希望総給与との差(絶対値)')

    write_sum(detail)

    wb.save(file_name)
    return

#月間の勤務表の書き込み
def write_monthly(sheet_name,staffs,days,opt_list):
    make_axis(sheet_name,staffs,days,'月間シフト')
    for i in staffs:
        for j in days:
            l = []
            for k in range(len(opt_list[i][j].index.values)):
                if opt_list[i][j].notnull()[opt_list[i][j].index.values[k]]:
                    if k == 0:
                        l.append(opt_list[i][j].index.values[k])
                    elif opt_list[i][j].isnull()[opt_list[i][j].index.values[k-1]]:
                        l.append(opt_list[i][j].index.values[k])
                    if k == len(opt_list[i][j].index.values)-1:
                        l.append(opt_list[i][j].index.values[k])
                    elif opt_list[i][j].isnull()[opt_list[i][j].index.values[k+1]]:
                        l.append(opt_list[i][j].index.values[k])
            count = 0
            time = ''
            for m in l:
                if count >= 2:
                    if count % 2 == 0:
                        time = time + '\n' + m.replace('.0',':00').replace('.5',':30') + '～'
                    else:
                        m = float(m) + 0.5
                        time = time + str(m).replace('.0',':00').replace('.5',':30')
                else:
                    if count % 2 == 0:
                        time = time + m.replace('.0',':00').replace('.5',':30') + '～'
                    else:
                        m = float(m) + 0.5
                        time = time + str(m).replace('.0',':00').replace('.5',':30')
                count += 1
            if time != '':
                sheet_name.cell(found_row(sheet_name,i),found_col(sheet_name,j),time).fill = PatternFill(patternType='lightHorizontal', fgColor = 'ffd700', bgColor = 'ffd700')
    adjust_width(sheet_name)
    return

#日別の勤務表の書き込み
def write_daily(wb,opt_list,days,staffs,slots):
    for j in [str(d) + '日' for d in days]:
        colors = make_colors(staffs,[p for p in opt_list[j].columns.values if p not in staffs])
        ws = wb.create_sheet(title = j)
        make_axis(ws,staffs,slots,'スタッフ別')
        add_axis(ws,[p for p in opt_list[j].columns.values if p not in staffs],slots,'ポジション別')
        for col_num in range(len(opt_list[j].columns.values)):
            for idx_num in range(len(opt_list[j].index.values)):
                col = opt_list[j].columns.values[col_num]
                if col in staffs:
                    plus = 2
                else:
                    plus = 4
                idx = opt_list[j].index.values[idx_num]
                if idx_num == 0:
                    if opt_list[j].notnull()[col][idx]:
                        ws.cell(col_num+plus,idx_num+2,opt_list[j][col][idx]).fill = PatternFill(patternType='lightHorizontal', fgColor=colors[opt_list[j][col][idx]], bgColor = colors[opt_list[j][col][idx]])
                        ws.cell(col_num+plus,idx_num+2).border = Border(top=thin, bottom=thin, left=thin)
                elif opt_list[j].isnull()[col][idx]:
                    if opt_list[j].notnull()[col][opt_list[j].index.values[idx_num-1]]:
                        ws.cell(col_num+plus,idx_num+2).border = Border(top=thin, bottom=thin, left=thin, right=thin)
                else:
                    if opt_list[j][col][idx] == opt_list[j][col][opt_list[j].index.values[idx_num-1]]:
                        ws.cell(col_num+plus,idx_num+2).fill = PatternFill(patternType='lightHorizontal', fgColor=colors[opt_list[j][col][idx]], bgColor = colors[opt_list[j][col][idx]])
                        ws.cell(col_num+plus,idx_num+2).border = Border(top=thin, bottom=thin)
                    else:
                        ws.cell(col_num+plus,idx_num+2,opt_list[j][col][idx]).fill = PatternFill(patternType='lightHorizontal', fgColor=colors[opt_list[j][col][idx]], bgColor = colors[opt_list[j][col][idx]])
                        ws.cell(col_num+plus,idx_num+2).border = Border(top=thin, bottom=thin, left=thin)
                if idx_num == len(opt_list[j].index.values)-1:
                    if opt_list[j].notnull()[col][idx]:
                        ws.cell(col_num+plus,idx_num+2).border = Border(top=thin, bottom=thin, right=thin)
                    else:
                        ws.cell(col_num+plus,idx_num+2).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        adjust_width(ws)
    return

#スタッフ別勤務表の書き込み
def write_staff(wb,opt_list,staffs,days,positions,slots):
    colors = make_colors(staffs,positions)
    for i in staffs:
        ws = wb.create_sheet(title = i)
        make_axis(ws,[j + '日' for j in days],slots,i)
        for col_num in range(len(opt_list[i].columns.values)):
            for idx_num in range(len(opt_list[i].index.values)):
                col = opt_list[i].columns.values[col_num]
                idx = opt_list[i].index.values[idx_num]
                if idx_num == 0:
                    if opt_list[i].notnull()[col][idx]:
                        ws.cell(col_num+2,idx_num+2,opt_list[i][col][idx]).fill = PatternFill(patternType='lightHorizontal', fgColor=colors[opt_list[i][col][idx]], bgColor = colors[opt_list[i][col][idx]])
                        ws.cell(col_num+2,idx_num+2).border = Border(top=thin, bottom=thin, left=thin)
                elif opt_list[i].isnull()[col][idx]:
                    if opt_list[i].notnull()[col][opt_list[i].index.values[idx_num-1]]:
                        ws.cell(col_num+2,idx_num+2).border = Border(top=thin, bottom=thin, left=thin, right=thin)
                else:
                    if opt_list[i][col][idx] == opt_list[i][col][opt_list[i].index.values[idx_num-1]]:
                        ws.cell(col_num+2,idx_num+2).fill = PatternFill(patternType='lightHorizontal', fgColor=colors[opt_list[i][col][idx]], bgColor = colors[opt_list[i][col][idx]])
                        ws.cell(col_num+2,idx_num+2).border = Border(top=thin, bottom=thin)
                    else:
                        ws.cell(col_num+2,idx_num+2,opt_list[i][col][idx]).fill = PatternFill(patternType='lightHorizontal', fgColor=colors[opt_list[i][col][idx]], bgColor = colors[opt_list[i][col][idx]])
                        ws.cell(col_num+2,idx_num+2).border = Border(top=thin, bottom=thin, left=thin)
                if idx_num == len(opt_list[i].index.values)-1:
                    if opt_list[i].notnull()[col][idx]:
                        ws.cell(col_num+2,idx_num+2).border = Border(top=thin, bottom=thin, right=thin)
                    else:
                        ws.cell(col_num+2,idx_num+2).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        adjust_width(ws)
    return

#詳細データの書き込み
def write_detail(sheet_name,d,chara):
    max_val = max(d.values())
    min_val = min(d.values())
    for k,v in d.items():
        if v == max_val:
            sheet_name.cell(found_row(sheet_name,k),found_col(sheet_name,chara),v).font = Font(bold=True, color='ff0000')
        elif v == min_val:
            sheet_name.cell(found_row(sheet_name,k),found_col(sheet_name,chara),v).font = Font(bold=True, color='0000ff')
        else:
            sheet_name.cell(found_row(sheet_name,k),found_col(sheet_name,chara),v).font = Font(bold=True)
        secure_width(sheet_name,found_col(sheet_name,chara),v)
    return

#総勤務時間を計算
def count_work_time(opt_list,days,staffs,positions,salarys):
    sum_work_time = dict(zip(staffs,[0]*len(staffs)))
    sum_day = dict(zip(staffs,[0]*len(staffs)))
    sum_night = dict(zip(staffs,[0]*len(staffs)))
    sum_salary = dict(zip(staffs,[0]*len(staffs)))
    staffs_posi = dict(zip(positions,[0]*len(positions)))
    for i in staffs:
        posi_count = dict(zip(positions,[0]*len(positions)))
        for j in days:
            posi_count_daily = dict(zip(positions,[0]*len(positions)))
            for k in opt_list[i][j].index.values:
                if opt_list[i].notnull()[j][k]:
                    sum_work_time[i] += 0.5
                    if int(float(k)) >= 22:
                        sum_night[i] += 0.5
                    else:
                        sum_day[i] += 0.5
                    posi_count_daily[opt_list[i][j][k]] = 1
            for l in positions:
                posi_count[l] += posi_count_daily[l]
        for l2 in positions:
            if staffs_posi[l2] == 0:
                staffs_posi[l2] = {i:posi_count[l2]} 
            else:
                staffs_posi[l2][i] = posi_count[l2]
        sum_salary[i] = int(sum_day[i] * salarys[i] + 1.25 * sum_night[i] * salarys[i])
    return sum_work_time,sum_day,sum_night,sum_salary,staffs_posi

#表の軸を作成
def make_axis(sheet_name,v_axis,h_axis,title):
    for i in range(len(v_axis)+1):
        for j in range(len(h_axis)+1):
            if i == 0 and j == 0:
                sheet_name.cell(1,1,title).border = Border(top=double, bottom=double, left=double, right=double)
                secure_width(sheet_name,1,title)
            elif i == 0:
                sheet_name.cell(1,j+1,h_axis[j-1]).border = Border(top=thin, bottom=double, right=thin)
                sheet_name.cell(1,j+1).alignment = Alignment(horizontal = 'center')
                secure_width(sheet_name,j+1,h_axis[j-1])
            elif j == 0:
                sheet_name.cell(i+1,1,v_axis[i-1]).border = Border(bottom=thin, left=thin, right=double)
                sheet_name.cell(i+1,1).alignment = Alignment(horizontal = 'center')
                secure_width(sheet_name,1,v_axis[i-1])
                if i % 2 == 1:
                    sheet_name.cell(i+1,1).fill = PatternFill(patternType='solid', fgColor='eeeeee', bgColor='eeeeee')
            else:
                sheet_name.cell(i+1,j+1).border = Border(bottom=thin, right=thin)
                sheet_name.cell(i+1,j+1).alignment = Alignment(horizontal = 'center')
                if i % 2 == 1:
                    sheet_name.cell(i+1,j+1).fill = PatternFill(patternType='solid', fgColor='eeeeee', bgColor='eeeeee')
    return

#表に新しい表を追加
def add_axis(sheet_name,v_axis,h_axis,title):
    row_idx = len(list(sheet_name.iter_rows()))
    for i in range(len(v_axis)+1):
        for j in range(len(h_axis)+1):
            if i == 0 and j == 0:
                sheet_name.cell(row_idx+2,1,title).border = Border(top=double, bottom=double, left=double, right=double)
                secure_width(sheet_name,1,title)
            elif i == 0:
                sheet_name.cell(row_idx+2,j+1,h_axis[j-1]).border = Border(top=thin, bottom=double, right=thin)
                sheet_name.cell(row_idx+2,j+1).alignment = Alignment(horizontal = 'center')
                secure_width(sheet_name,j+1,h_axis[j-1])
            elif j == 0:
                sheet_name.cell(row_idx+i+2,1,v_axis[i-1]).border = Border(bottom=thin, left=thin, right=double)
                sheet_name.cell(row_idx+i+2,1).alignment = Alignment(horizontal = 'center')
                secure_width(sheet_name,1,v_axis[i-1])
                if i % 2 == 1:
                    sheet_name.cell(row_idx+i+2,1).fill = PatternFill(patternType='solid', fgColor='eeeeee', bgColor='eeeeee')
            else:
                sheet_name.cell(row_idx+i+2,j+1).border = Border(bottom=thin, right=thin)
                sheet_name.cell(row_idx+i+2,j+1).alignment = Alignment(horizontal = 'center')
                if i % 2 == 1:
                    sheet_name.cell(row_idx+i+2,j+1).fill = PatternFill(patternType='solid', fgColor='eeeeee', bgColor='eeeeee')
    return

#列の値を変更
def change_h_axis(sheet,h_axis,row):
    for i in range(len(h_axis)):
        sheet.cell(row,i+2,h_axis[i]).border = Border(top=thin, bottom=double, right=thin)
    return    

#勤務日数を計算
def count_work_day(sheet_name,staffs):
    sum_work_day = {}
    sum_off = {}
    max_conti = {}
    for i in staffs:
        count_work = 0
        count_off = 0
        conti = 0
        max_conti[i] = 0
        idx = found_row(sheet_name,i)
        for col in sheet_name.iter_cols(min_col=2):
            if col[idx-1].value != None:
                count_work += 1
                conti += 1
                if conti > max_conti[i]:
                    max_conti[i] = conti
            else:
                count_off += 1
                conti = 0
        sum_work_day[i] = count_work
        sum_off[i] = count_off
    return sum_work_day,sum_off,max_conti

#詳細データの合計を計算
def write_sum(sheet_name):
    idx_row = found_row(sheet_name,'合計')
    for i in list(sheet_name.rows)[0]:
        if i.column != 1:
            sum_v = 0
            for j in list(sheet_name.columns)[i.column-1]:
                if type(j.value) == int or type(j.value) == float:
                    sum_v += j.value
            sheet_name.cell(idx_row,i.column,sum_v).font = Font(bold=True)
            secure_width(sheet_name,i.column,sum_v)
    return

#セルの幅を値毎に揃える
def adjust_width(sheet_name):
    for col in sheet_name.iter_cols():
        max_len = 0
        column = col[0].column
        for cell in col:
            if count_len(str(cell.value)) > max_len:
                max_len = count_len(str(cell.value))
            cell.alignment = openpyxl.styles.Alignment(horizontal = 'center',vertical = 'center',wrapText=True)
        adjusted_width = (max_len + 2)
        sheet_name.column_dimensions[get_column_letter(column)].width = adjusted_width
    return

#セルの幅を確保する
def secure_width(sheet_name,col,value):
    if count_len(str(value)) > sheet_name.column_dimensions[get_column_letter(col)].width:
        sheet_name.column_dimensions[get_column_letter(col)].width = count_len(str(value)) 
    return

#テキストの長さを計算
def count_len(text):
    count = 0
    for i in text:
        if unicodedata.east_asian_width(i) in 'FWA':
            count += 2
        else:
            count += 1
    return count

#スタッフとポジションの色を指定
def make_colors(staffs,positions):
    max_rgb,r,g,b, = 255,0,0,0
    colors = {i:'ffffff' for i in staffs + positions}
    color_len = len(colors)
    inter = math.floor(max_rgb / math.ceil(color_len/6))
    count = 0
    for k in colors.keys():
        if count % 6 == 0:
            cellcolor = '%02x%02x%02x' % (max_rgb,g,b)
        if count % 6 == 1:
            cellcolor = '%02x%02x%02x' % (r,max_rgb,b)
        if count % 6 == 2:
            cellcolor = '%02x%02x%02x' % (r,g,max_rgb)
        if count % 6 == 3:
            cellcolor = '%02x%02x%02x' % (max_rgb,max_rgb,b)
        if count % 6 == 4:
            cellcolor = '%02x%02x%02x' % (max_rgb,g,max_rgb)
        if count % 6 == 5:
            cellcolor = '%02x%02x%02x' % (r,max_rgb,max_rgb)
            r += inter
            g += inter
            b += inter
        count += 1
        colors[k] = cellcolor
    return colors

if __name__ == '__main__':
    main()